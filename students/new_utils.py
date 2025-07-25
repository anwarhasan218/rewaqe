import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Student, StudentResult, Level, Course, Stage
from django.utils.translation import gettext_lazy as _
import csv
from openpyxl.utils import get_column_letter

def export_comprehensive_results_template(level, academic_year=None):
    """تصدير نموذج النتائج الشاملة بصيغة Excel"""

    # إنشاء دفتر عمل جديد
    wb = Workbook()
    ws = wb.active
    ws.title = f"نموذج نتائج {level.stage.name} - {level.name}"

    # تعريف الألوان والأنماط
    header_fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
    info_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    header_font = Font(bold=True, color='FFFFFF')

    # معلومات المستوى
    ws['A1'] = f"نموذج نتائج {level.stage.name} - {level.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = info_fill
    ws.merge_cells('A1:D1')

    ws['A2'] = "تعليمات:"
    ws['A2'].font = Font(bold=True)
    ws['A3'] = "1. املأ درجات جميع المواد لكل طالب"
    ws['A4'] = "2. لا تغير أسماء الأعمدة أو أكواد الطلاب"
    ws['A5'] = "3. احفظ الملف بصيغة Excel (.xlsx)"

    # إنشاء العناوين
    headers = ['رقم الجلوس', 'اسم الطالب']
    courses = Course.objects.filter(level=level).order_by('name')

    for course in courses:
        headers.append(f'{course.name}\n(من {course.max_score})')

    # كتابة العناوين في الصف 7
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # إضافة بيانات الطلاب
    students = Student.objects.filter(level=level).order_by('current_seat_number')

    for row_num, student in enumerate(students, 8):
        ws.cell(row=row_num, column=1, value=student.current_seat_number)
        ws.cell(row=row_num, column=2, value=student.full_name)

        # إضافة خانات فارغة للدرجات
        for col in range(3, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col, value='')
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    # ضبط عرض الأعمدة
    for i, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    # إنشاء الاستجابة
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"comprehensive_results_template_{level.stage.name}_{level.name}.xlsx"
    filename = filename.replace(' ', '_').replace('و', 'w')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def export_comprehensive_results(level, batch_id=None):
    """تصدير النتائج الشاملة بصيغة Excel"""

    # إنشاء دفتر عمل جديد
    wb = Workbook()
    ws = wb.active
    ws.title = f"نتائج {level.stage.name} - {level.name}"

    # تعريف الألوان والأنماط
    header_fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    header_font = Font(bold=True, color='FFFFFF')

    # إنشاء العناوين
    headers = ['الكود', 'رقم الجلوس', 'الاسم']
    courses = Course.objects.filter(level=level).order_by('name')

    for course in courses:
        headers.append(course.name)

    headers.extend(['المجموع', 'النسبة %', 'النتيجة', 'مواد الرسوب'])

    # كتابة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # الحصول على النتائج
    results_filter = {'level': level}
    if batch_id:
        results_filter['batch_id'] = batch_id

    results = StudentResult.objects.filter(**results_filter).order_by('student__code')

    # كتابة بيانات الطلاب
    for row_num, result in enumerate(results, 2):
        ws.cell(row=row_num, column=1, value=result.student.code)
        ws.cell(row=row_num, column=2, value=result.student.current_seat_number or '')
        ws.cell(row=row_num, column=3, value=result.student.full_name)

        col_num = 4
        # إضافة درجات المواد
        for course in courses:
            score = result.get_course_score(course.id)
            ws.cell(row=row_num, column=col_num, value=score if score is not None else 'غ')
            col_num += 1

        # إضافة المجموع والنتيجة
        ws.cell(row=row_num, column=col_num, value=result.total_score)
        ws.cell(row=row_num, column=col_num + 1, value=result.percentage)
        ws.cell(row=row_num, column=col_num + 2, value=result.result)

        # إضافة مواد الرسوب
        failed_courses = result.failed_courses.all()
        failed_text = ', '.join([course.name for course in failed_courses]) if failed_courses else ''
        ws.cell(row=row_num, column=col_num + 3, value=failed_text)

    # ضبط عرض الأعمدة
    for i, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    # إنشاء الاستجابة
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"comprehensive_results_{level.stage.name}_{level.name}"
    if batch_id:
        from .models import Batch
        try:
            batch = Batch.objects.get(id=batch_id)
            filename += f"_{batch.name}"
        except:
            pass
    filename += ".xlsx"
    filename = filename.replace(' ', '_').replace('و', 'w')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def export_student_comprehensive_result(student, academic_year=None):
    """تصدير بيان درجات طالب واحد (نظام شامل)"""

    # إنشاء دفتر عمل جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "بيان درجات"

    # تعريف الألوان والأنماط
    header_fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
    pass_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    # إعداد معلومات الطالب
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "بيان درجات الطالب"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = header_fill

    # معلومات الطالب
    student_info = [
        ['اسم الطالب:', student.full_name],
        ['كود الطالب:', student.code],
        ['المستوى:', student.level.name],
        ['المرحلة:', student.level.stage.name],
        ['العام الدراسي:', academic_year or 'جميع السنوات']
    ]

    for i, (label, value) in enumerate(student_info, 3):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)

    # الحصول على نتائج الطالب
    if academic_year:
        student_results = StudentResult.objects.filter(student=student, academic_year=academic_year).first()
    else:
        student_results = StudentResult.objects.filter(student=student).order_by('-academic_year').first()

    if student_results:
        # إضافة جدول النتائج
        headers = ['م', 'المادة', 'الدرجة', 'النهاية العظمى', 'الحالة']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # إضافة درجات المواد
        courses = Course.objects.filter(level=student.level).order_by('name')
        row_num = 10

        for i, course in enumerate(courses, 1):
            score = student_results.get_course_score(course.id)
            if score is not None:
                if isinstance(score, int):
                    is_pass = score >= course.pass_score
                    status = "ناجح" if is_pass else "راسب"
                    fill = pass_fill if is_pass else fail_fill
                else: # غائب
                    status = "غائب"
                    fill = fail_fill
            else:
                score = "N/A"
                status = "N/A"
                fill = PatternFill()


            ws.cell(row=row_num, column=1, value=i).border = border
            ws.cell(row=row_num, column=2, value=course.name).border = border
            ws.cell(row=row_num, column=3, value=score).border = border
            ws.cell(row=row_num, column=4, value=course.max_score).border = border
            status_cell = ws.cell(row=row_num, column=5, value=status)
            status_cell.border = border
            status_cell.fill = fill
            row_num += 1

        # إضافة ملخص النتيجة
        summary_start_row = row_num + 2
        ws.cell(row=summary_start_row, column=1, value="المجموع الكلي:").font = Font(bold=True)
        ws.cell(row=summary_start_row, column=2, value=student_results.total_score)
        ws.cell(row=summary_start_row+1, column=1, value="النسبة المئوية:").font = Font(bold=True)
        ws.cell(row=summary_start_row+1, column=2, value=f"{student_results.percentage:.2f}%")
        ws.cell(row=summary_start_row+2, column=1, value="النتيجة النهائية:").font = Font(bold=True)
        ws.cell(row=summary_start_row+2, column=2, value=student_results.get_result_display())
        ws.cell(row=summary_start_row+3, column=1, value="تاريخ النتيجة:").font = Font(bold=True)
        ws.cell(row=summary_start_row+3, column=2, value=student_results.result_date.strftime('%Y-%m-%d'))


    # ضبط عرض الأعمدة
    for i, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"student_result_{student.code}_{academic_year or 'latest'}.xlsx"
    filename = filename.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response



def export_application_terms(level=None):
    """
    تصدير مصطلحات التقديم (المواد والمستويات) إلى ملف Excel.
    إذا تم تحديد مستوى، يتم تصدير المواد الخاصة به فقط.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "مصطلحات التقديم"

    # الأنماط
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # كتابة العناوين
    headers = ["المرحلة", "المستوى", "المادة", "الدرجة العظمى", "درجة النجاح"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # كتابة البيانات
    row = 2
    if level:
        courses = Course.objects.filter(level=level).order_by('name')
        if not courses:
            ws.cell(row, 1, f"لا توجد مواد دراسية للمستوى المحدد: {level.name}")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    else:
        courses = Course.objects.all().order_by('level__stage__name', 'level__name', 'name')

    for course in courses:
        ws.cell(row, 1, course.level.stage.name).border = border
        ws.cell(row, 2, course.level.name).border = border
        ws.cell(row, 3, course.name).border = border
        ws.cell(row, 4, course.max_score).border = border
        ws.cell(row, 5, course.pass_score).border = border
        row += 1

    # ضبط عرض الأعمدة
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"application_terms_{level.name if level else 'all'}.xlsx"
    filename = filename.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
